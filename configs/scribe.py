import csv
import os
import datetime
import numpy
import pathlib
# import xlsxwriter
import openpyxl

class Scribe:
    def __init__(self, project):
        self.current_file_line_counter = 0
        self.file_counter = 1
        self.test_dir = os.getcwd()
        self.data_dir = os.path.join(self.test_dir, f"{project}_data")
        self.base_dir = os.path.join(self.test_dir, f"{project}_data")

        self.headers = []

    def create_session_dir(self):
        start_timestamp = datetime.datetime.now()
        session = start_timestamp.strftime("%Y%m%d%H%M%S")
        self.data_dir = os.path.join(self.data_dir, session)
        if not os.path.exists(self.data_dir):
            os.mkdir(self.data_dir)

    def change_sno_dir(self, sno):
        self.data_dir = os.path.join(self.data_dir, sno)
        if not os.path.exists(self.data_dir):
            os.mkdir(self.data_dir)

    def change_results_dir(self, results_dir_name):
        self.data_dir = os.path.join(self.data_dir, results_dir_name)
        if not os.path.exists(self.data_dir):
            os.mkdir(self.data_dir)

    def new_sno(self, sno, results_dir_name):
        self.data_dir = self.base_dir
        self.change_sno_dir(sno)
        self.change_results_dir(results_dir_name)
        self.create_session_dir()
        self.init_fnames()

    def init_fnames(self):
        self.bandwidth_module_fnames = {
            'LOW_BAND_PATH1_20DB': os.path.join(self.data_dir, "LOW_BAND_PATH1_20DB_BANDWIDTH_DATA.csv"),
            'LOW_BAND_PATH2_20DB': os.path.join(self.data_dir, "LOW_BAND_PATH2_20DB_BANDWIDTH_DATA.csv"),
            'HIGH_BAND_PATH1_20DB': os.path.join(self.data_dir, "HIGH_BAND_PATH1_20DB_BANDWIDTH_DATA.csv"),
            'HIGH_BAND_PATH2_20DB': os.path.join(self.data_dir, "HIGH_BAND_PATH2_20DB_BANDWIDTH_DATA.csv"),
        }

        self.power_meter_module_fnames = {
            'LOW_BAND_PATH1_20DB': os.path.join(self.data_dir, "LOW_BAND_PATH1_20DB_POWER_METER_DATA.csv"),
            'LOW_BAND_PATH2_20DB': os.path.join(self.data_dir, "LOW_BAND_PATH2_20DB_POWER_METER_DATA.csv"),
            'HIGH_BAND_PATH1_20DB': os.path.join(self.data_dir, "HIGH_BAND_PATH1_20DB_POWER_METER_DATA.csv"),
            'HIGH_BAND_PATH2_20DB': os.path.join(self.data_dir, "HIGH_BAND_PATH2_20DB_POWER_METER_DATA.csv"),
        }

        self.na_21_module_fnames = {
            'LOW_BAND_PATH1_20DB': [os.path.join(self.data_dir, "LOW_BAND_PATH1_20DB_S21_MLOG_DATA.csv"), os.path.join(self.data_dir, "LOW_BAND_PATH1_20DB_S21_PHASE_DATA.csv")],
            'LOW_BAND_PATH2_20DB': [os.path.join(self.data_dir, "LOW_BAND_PATH2_20DB_S21_MLOG_DATA.csv"), os.path.join(self.data_dir, "LOW_BAND_PATH2_20DB_S21_PHASE_DATA.csv")],
            'HIGH_BAND_PATH1_20DB': [os.path.join(self.data_dir, "HIGH_BAND_PATH1_20DB_S21_MLOG_DATA.csv"), os.path.join(self.data_dir, "HIGH_BAND_PATH1_20DB_S21_PHASE_DATA.csv")],
            'HIGH_BAND_PATH2_20DB': [os.path.join(self.data_dir, "HIGH_BAND_PATH2_20DB_S21_MLOG_DATA.csv"), os.path.join(self.data_dir, "HIGH_BAND_PATH2_20DB_S21_PHASE_DATA.csv")]
        }

        self.na_22_module_fnames = {
            'LOW_BAND_PATH1_10DB': [os.path.join(self.data_dir, "LOW_BAND_PATH1_10DB_S22_MLOG_DATA.csv"), os.path.join(self.data_dir, "LOW_BAND_PATH1_10DB_S22_PHASE_DATA.csv")],
            'LOW_BAND_PATH2_10DB': [os.path.join(self.data_dir, "LOW_BAND_PATH2_10DB_S22_MLOG_DATA.csv"), os.path.join(self.data_dir, "LOW_BAND_PATH2_10DB_S22_PHASE_DATA.csv")],
            "HIGH_BAND_PATH1_10DB": [os.path.join(self.data_dir, "HIGH_BAND_PATH1_10DB_S22_MLOG_DATA.csv"), os.path.join(self.data_dir, "HIGH_BAND_PATH1_10DB_S22_PHASE_DATA.csv")],
            "HIGH_BAND_PATH2_10DB": [os.path.join(self.data_dir, "HIGH_BAND_PATH2_10DB_S22_MLOG_DATA.csv"), os.path.join(self.data_dir, "HIGH_BAND_PATH2_10DB_S22_PHASE_DATA.csv")]
        }

        self.na_11_module_fnames = {
            'LOW_BAND_PATH1_20DB': [os.path.join(self.data_dir, "LOW_BAND_PATH1_20DB_S11_MLOG_DATA.csv"), os.path.join(self.data_dir, "LOW_BAND_PATH1_20DB_S11_PHASE_DATA.csv")],
            'LOW_BAND_PATH2_20DB': [os.path.join(self.data_dir, "LOW_BAND_PATH2_20DB_S11_MLOG_DATA.csv"), os.path.join(self.data_dir, "LOW_BAND_PATH2_20DB_S11_PHASE_DATA.csv")],
            'HIGH_BAND_PATH1_20DB': [os.path.join(self.data_dir, "HIGH_BAND_PATH1_20DB_S11_MLOG_DATA.csv"), os.path.join(self.data_dir, "HIGH_BAND_PATH1_20DB_S11_PHASE_DATA.csv")],
            'HIGH_BAND_PATH2_20DB': [os.path.join(self.data_dir, "HIGH_BAND_PATH2_20DB_S11_MLOG_DATA.csv"), os.path.join(self.data_dir, "HIGH_BAND_PATH2_20DB_S11_PHASE_DATA.csv")]
        }

    
    def write_data_from_filepath(self, filepath, data):
        print(filepath)
        with open(filepath, 'a', encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(data)
            f.close()

    def write_bandwidth_data_from_array(self, switchpath, data_array):
        with open(self.bandwidth_fnames[switchpath], 'a', encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(data_array)
            f.close()

    def write_bandwidth_module_data_from_array(self, switchpath, data_array):
        with open(self.bandwidth_module_fnames[switchpath], 'a', encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(data_array)
            f.close()


    def write_power_meter_module_data_from_array(self, switchpath, data_array):
        with open(self.power_meter_module_fnames[switchpath], 'a', encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(data_array)
            f.close()

    def write_power_meter_data_from_array(self, switchpath, data_array):
        with open(self.power_meter_fnames[switchpath], 'a', encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(data_array)
            f.close()

    def write_na_module_data(self, switchpath, ratioed_power, format, data):
        bat = {}
        if ratioed_power == "S22":
            bat = self.na_22_module_fnames
        elif ratioed_power == "S21":
            bat = self.na_21_module_fnames
        elif ratioed_power == "S11":
            bat = self.na_11_module_fnames

        bat = bat[switchpath]

        if format == "MLOG":
            fname = bat[0]
        elif format == "PHASE":
            fname = bat[1]

        with open(fname, 'a', encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(data)
            f.close()
        

    def write_na_gain_data_from_array(self, switchpath, data_array):
        with open(self.na_gain_fnames[switchpath], 'a', encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(data_array)
            f.close()

    def write_na_phase_data_from_array(self, switchpath, data_array):
        with open(self.na_phase_fnames[switchpath], 'a', encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(data_array)
            f.close()

    def write_data_from_array_to_new_column(self, ):
        print("SHO")

    def create_new_csv_file(self):
        with open(self.current_dir + self.current_fname, "w", encoding="utf-8") as f:
            f.write(self.headers)
            f.close()

    def reset_line_counter(self):
        self.current_file_line_counter = 0

    def csv_to_xlsx(self, csv_filepath, xlsx_filepath, test_type):
        # Transpose data from rows to columns
        with open(csv_filepath, 'r') as f:
            reader = csv.reader(f)
            data = list(reader)

        # Check the test type and transpose the data accordingly
        if test_type == "bandwidth":
            data_bucket = numpy.array(data).T.tolist()

        # Write the transposed data to a new xlsx file using openpyxl
        workbook = openpyxl.Workbook(xlsx_filepath)
        workbook.create_sheet("bandwidth")
        sheet = workbook["bandwidth"]
        print(data_bucket)
        for row in data_bucket:
            sheet.append(row)

        workbook.save(xlsx_filepath)

    def append_transposed_data_to_xlsx(self, data, xlsx_filepath, sheet_name):
        # Receieve data transposed from csv_to_xlsx
        workbook = openpyxl.load_workbook(xlsx_filepath)
        sheet = workbook[sheet_name]
        max_col = sheet.max_column
        max_row = sheet.max_row

        # for sheet_row in sheet.iter_rows(min_col=max_col + 1, max_col=max_col * 2, max_row=max_row):
        #     data_bucket[]
        #     for i, cell in enumerate(sheet_row):
        #         cell.value = row[i]

        # workbook.save(xlsx_filepath)

    def get_p1_data(self, linear_gain_bucket, sat_gain_bucket):
        base = os.getcwd()
        filepath = os.path.join(base, "p1.csv")

        num_of_points = len(linear_gain_bucket[0])
        freqs = linear_gain_bucket[0][1::]
        linear_gain_bucket = linear_gain_bucket[1::]
        linear_gain_bucket_len = len(linear_gain_bucket)

        sum_buck = []
        for i in range(0, num_of_points):
            sum_buck.append(0)

        for i, linear_gain in enumerate(linear_gain_bucket):
            for j, gain_value in enumerate(linear_gain):
                if  j == 0:
                    continue
                else:
                    sum_buck[j] += gain_value

        sum_buck = sum_buck[1::]

        avg_linear_line = []
        for index, sum in enumerate(sum_buck):
            avg = sum / linear_gain_bucket_len
            avg_linear_line.append(avg)

        print(f"AVERAGE LIN LINE: {len(avg_linear_line)}")
        p1_data = {}

        for freq in freqs:
            p1_data[freq] = 0

        for gain_saturated in sat_gain_bucket:
            gain_buck = gain_saturated[1::]
            for index, gain_value in enumerate(gain_buck):
                if p1_data[freqs[index]] == 0:
                    diff = gain_value - avg_linear_line[index]
                    print(diff)
                    if diff < -1:
                        p1_data[freqs[index]] = gain_saturated[0]


        with open(filepath, mode="a", newline="") as f:
            writer = csv.writer(f)
            for frequency, source_db in p1_data.items():
                writer.writerow([frequency, source_db])

            f.close()

    def get_p1_data_v2(self, freqs, linear_gain_bucket, saturation_gain_bucket):
        source_dbs = linear_gain_bucket.keys()

        avg_line = [0 for i in range(0, len(freqs))]

        if isinstance(linear_gain_bucket, dict):
            linear_gain_bucket = linear_gain_bucket
        else:
            raise TypeError("linear_gain_bucket must be a dictionary")
        

        if isinstance(saturation_gain_bucket, dict):
            saturation_gain_bucket = saturation_gain_bucket
        else:
            raise TypeError("saturation_gain_bucket must be a dictionary")

        for source_db, data in linear_gain_bucket.items():
            for i, gain_value in enumerate(data["gain"]):
                avg_line[i] += gain_value
        
        max_linear = max(source_dbs)
        min_linear = min(source_dbs)
        
        for i, avg_sum in enumerate(avg_line):
            avg_line[i] = avg_sum / len(source_dbs)

        print(avg_line)

        p1_data = {}
        psat_data = {}

        for i, freq in enumerate(freqs):
            for source_db, data in saturation_gain_bucket.items():
                if freq not in p1_data:
                    if data["gain"][i] - avg_line[i] < -1:
                        p1_data[freq] = source_db
        
        for i, freq in enumerate(freqs):
            psat_data[freq] = []

        for source_db, data in saturation_gain_bucket.items():
            for i, gain_value in enumerate(data["gain"]):
                psat_data[freqs[i]].append(gain_value + source_db)

        for freq, array in psat_data.items():
            psat_data[freq] = max(array)
            
        return p1_data, psat_data