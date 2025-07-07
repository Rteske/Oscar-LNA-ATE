# Take data from csv file and convert it to a list of dictionaries
import csv
from typing import List, Dict
from datetime import datetime
import os
import json
import logging
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook   
import openpyxl
import shutil
import pandas as pd
import numpy as np

import matplotlib.pyplot as plt
from matplotlib.ticker import PercentFormatter

class TestDataSheetWriter:
    """
    Class to represent a test data sheet.
    """
    def __init__(self):
        """
        Initialize the TestDataSheetWriter class.
        """
        # Initialize the base filepaths
        self.base_csv_filepath = os.path.join(os.path.dirname(__file__), 'OSCAR_practice')
        self.base_csv_filepath = os.path.abspath(self.base_csv_filepath)

        # Initialize the base templates xlsx filepath
        self.base_xlsx_filepath = os.path.join(os.path.dirname(__file__), "OSCAR TEMPLATES")
        self.base_xlsx_filepath = os.path.abspath(self.base_xlsx_filepath)

        # Initialize the base tds filepath
        datetime_dir = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        self.base_tds_filepath = os.path.join(os.path.dirname(__file__), 'OSCAR_PA_TDS')
        self.base_tds_filepath = os.path.join(self.base_tds_filepath, datetime_dir)
        # Create the directory if it doesn't exist
        if not os.path.exists(self.base_tds_filepath):
            os.makedirs(self.base_tds_filepath)
        self.base_tds_filepath = os.path.abspath(self.base_tds_filepath)


        # Create a permanent lookup table for what data to extract and where it goes
        # xlsx file are the keys and the values are the csv file names
        self.tds_lookup = {
            "LOW_BAND_PATH1_20DB_BANDWIDTH_DATA": {
                "filepaths": [
                    "Low Band 10 dB Atten OQPSK plots.xlsx",
                    "Low Band max gain OQPSK plots.xlsx",
                    "Low Band OQPSK 200 MHz Span plots.xlsx",
                    "Low Band Wideband Plots.xlsx"
                ],
            },
            "LOW_BAND_PATH2_20DB_BANDWIDTH_DATA": {
                "filepaths": [
                    "Low Band 10 dB Atten OQPSK plots.xlsx",
                    "Low Band max gain OQPSK plots.xlsx",
                    "Low Band OQPSK 200 MHz Span plots.xlsx",
                    "Low Band Wideband Plots.xlsx"
                ]
            },
            "HIGH_BAND_PATH1_20DB_BANDWIDTH_DATA": {
                "filepaths": [
                    "High Band 10 dB Atten OQPSK plots.xlsx",
                    "High Band max gain OQPSK plots.xlsx",
                    "High Band OQPSK 200 MHz Span plots.xlsx",
                    "High Band Wideband Plots.xlsx"
                ]
            },
            "HIGH_BAND_PATH2_20DB_BANDWIDTH_DATA": {
                "filepaths": [
                    "High Band 10 dB Atten OQPSK plots.xlsx",
                    "High Band max gain OQPSK plots.xlsx",
                    "High Band OQPSK 200 MHz Span plots.xlsx",
                    "High Band Wideband Plots.xlsx"
                ],
            },
            "LOW_BAND_PATH1_20DB_S21_MLOG_DATA": {
                "filepaths": "Low Band gain_phase_vswr",
            },
            "LOW_BAND_PATH2_20DB_S21_MLOG_DATA": {
                "filepaths": "Low Band gain_phase_vswr",
            },
            "HIGH_BAND_PATH1_20DB_S21_MLOG_DATA": {
                "filepaths": "High Band gain_phase_vswr",
            },
            "HIGH_BAND_PATH2_20DB_S21_MLOG_DATA": {
                "filepaths": "High Band gain_phase_vswr",
            },
            "LOW_BAND_PATH1_20DB_S21_PHASE_DATA": {
                "filepaths": "Low Band gain_phase_vswr",
            },
            "LOW_BAND_PATH2_20DB_S21_PHASE_DATA": {
                "filepaths": "Low Band gain_phase_vswr",
            },
            "HIGH_BAND_PATH1_20DB_S21_PHASE_DATA": {
                "filepaths": "High Band gain_phase_vswr",
            },
            "HIGH_BAND_PATH2_20DB_S21_PHASE_DATA": {
                "filepaths": "High Band gain_phase_vswr",
            },
            "LOW_BAND_PATH1_20DB_S11_MLOG_DATA": {
                "filepaths": "Low Band gain_phase_vswr",
            },
            "LOW_BAND_PATH2_20DB_S11_MLOG_DATA": {
                "filepaths": "Low Band gain_phase_vswr",
            },
            "HIGH_BAND_PATH1_20DB_S11_MLOG_DATA": {
                "filepaths": "High Band gain_phase_vswr",
            },
            "HIGH_BAND_PATH2_20DB_S11_MLOG_DATA": {
                "filepaths": "High Band gain_phase_vswr",
            },
            "LOW_BAND_PATH1_10DB_S22_MLOG_DATA": {
                "filepaths": "Low Band gain_phase_vswr",
            },
            "LOW_BAND_PATH2_10DB_S22_MLOG_DATA": {
                "filepaths": "Low Band gain_phase_vswr",
            },
            "HIGH_BAND_PATH1_10DB_S22_MLOG_DATA": {
                "filepaths": "High Band gain_phase_vswr",
            },
            "HIGH_BAND_PATH2_10DB_S22_MLOG_DATA": {
                "filepaths": "High Band gain_phase_vswr",
            },
        }
    
    def get_xlsx_filepaths(self, csv_filename: str) -> List[str]:
        """
        Get the xlsx filepaths from the lookup table based on the csv filename.
        """

        csv_filename = os.path.basename(csv_filename)
        csv_filename = csv_filename.replace(".csv", "")
        filepaths = []
        for key, value in self.tds_lookup.items():
            if csv_filename in key:
                filepaths = value.get("filepaths", [])
                break
        
        return filepaths

    # File dialog using tkinter
    def get_data_dir(self):
        """
        Open a file dialog to select a directory.
        """
        filepath = filedialog.askdirectory(initialdir=self.base_csv_filepath ,title="Select a directory containing data files")
        if not filepath:
            raise ValueError("No directory selected")
        
        # Check if the selected directory contains CSV files
        if not any(file.endswith('.csv') for file in os.listdir(filepath)):
            raise ValueError("No CSV files found in the selected directory")
        
        return filepath
    
    def get_csv_filepaths(self, directory: str) -> List[str]: 
        """
        Get a list of CSV files in the specified directory.
        
        """
        csv_filepaths = []
        for file in os.listdir(directory):
            if file.endswith('.csv'):
                csv_filepaths.append(os.path.join(directory, file))

        return csv_filepaths
    
    def read_csv_file(self, filepath: str) -> List:
        """
        Read a CSV file and return its contents as a list of dictionaries.
        """
        with open(filepath, 'r') as csvfile:
            csv_reader = csv.reader(csvfile)
            print(f"Reading file: {filepath}")
            bucket = []
            for row in csv_reader:
                mini_bucket = []
                for i, value in enumerate(row):
                    # Check if the value is a number
                    try:
                        value = float(value)
                    except ValueError:
                        pass
                    mini_bucket.append(value)

                bucket.append(mini_bucket)
        
        return bucket
    
    def ask_user_for_write_dir(self):
        self.base_tds_filepath = filedialog.askdirectory(initialdir=self.base_tds_filepath, title="Select a directory to write the test data sheets")

        if not self.base_tds_filepath:
            raise ValueError("No directory selected for writing test data sheets")
        

    def create_10mghz_plot(self, traces: List, destination_dir: str, gain_step: int, spec_line: List):
        """
        Create and write bandwidth measurement plots
        """

        step = 0.02
        shit = []
        for num in range(len(traces[0]['data'])):
            shit.append(num * step)

        plt.figure(figsize=(10, 6))
        plt.title(f"{gain_step} dB Attenuation, 10 MHz Span")
        plt.xlabel("Frequency (Hz)")
        plt.ylabel("Magnitude (dB)")
        plt.grid(True)
        plt.xscale('log')
        # Format xaxis as percentage of nessacary bandwidth
        plt.gca().xaxis.set_major_formatter(PercentFormatter())
        plt.ylim(-100, 0)  # Set y-axis limits for better visibility

        shit = shit[:round(len(traces[0]['data']) / 2)]
        percentages = shit[::-1]

        if len(percentages) > len(spec_line):
            spec_percentages = percentages[:len(spec_line)]


        plt.plot(spec_percentages[::-1], spec_line)
        for trace in traces:
            data = trace['data'][:round(len(trace['data']) / 2)]
            plt.plot(percentages, data, label=f"{trace['freq']}")

        plt.legend()

        plt.savefig(os.path.join(destination_dir, f"{gain_step}db_10mghz_plot.png"))
        plt.close()

    def create_200mghz_plot(self, traces: List, destination_dir: str, gain_step: int):
        plt.figure(figsize=(10, 6))
        plt.title(f"200 MHz Span")
        plt.xlabel("Frequency (Hz)")
        plt.ylabel("Magnitude (dB)")
        plt.grid(True)
        # Format xaxis as percentage of nessacary bandwidth
        plt.ylim(-100, 0)  # Set y-axis limits for better visibility
        for trace in traces:
            data = trace['data']
            plt.plot(data, label=f"{trace['freq']}")
        plt.legend()

        plt.savefig(os.path.join(destination_dir, f"{gain_step}db_200mghz_plot.png"))
        plt.close()

    def create_wideband_plot(self, traces: List, destination_dir: str, gain_step: int):
        """
        Create and write wideband measurement plots
        """
        plt.figure(figsize=(10, 6))
        plt.title(f"Wideband Span")
        plt.xlabel("Frequency (Hz)")
        plt.ylabel("Magnitude (dB)")
        plt.grid(True)
        # Format xaxis as percentage of nessacary bandwidth
        plt.ylim(-100, 0)

        for trace in traces:
            data = trace['data']
            frequencies = trace['frequencies']
            plt.plot(frequencies, data, label=f"{trace['attenuation_setting']} dB")

        plt.legend()
        plt.savefig(os.path.join(destination_dir, f"{gain_step}db_wideband_plot.png"))
        plt.close()

    def write_bandwidth_tds(self, data: List, csv_filepath: str, temp: str):
        """
        Write spectral mask data to an xlsx file.
        """

        if temp == "25c":
            temp_coords_offset = 0
        elif temp == "-13c":
            temp_coords_offset = 6
            if "HIGH_BAND" in csv_filepath:
                temp_coords_offset += 6
        elif temp == "55c":
            temp_coords_offset = 3
            if "HIGH_BAND" in csv_filepath:
                temp_coords_offset += 6
        
        df = pd.read_csv(csv_filepath, header=None)
        traces = []
        for i, row in df.iterrows():
            if i % 2 == 0:
                # Odd rows are the headers
                trace = {}
                headers = row.tolist()
                metadata = headers[:9]  # First 9 columns are metadata
                frequencies = headers[9:]  # Remaining columns are frequencies for the trace data

                for j, header in enumerate(metadata):
                    trace[header] = df.loc[i + 1].tolist()[j]

                trace['frequencies'] = frequencies
                trace['data'] = df.loc[i + 1].tolist()[9:]  # Get the trace data
                traces.append(trace)
            else:
                continue


        xlsx_filepaths = self.get_xlsx_filepaths(csv_filepath)

        
        max_gain_spec_line = [
            0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
            -7, -7.967950853, -8.649284425, -9.304901072, -9.93667176, -10.54627042, -11.13520069,
            -11.70481826, -12.25634964, -12.7909081, -13.30950719, -13.81307228, -14.30245054,
            -14.77841947, -15.24169437, -15.69293479, -16.13275016, -16.56170477, -16.98032212,
            -17.38908873, -17.78845757, -18.17885106, -18.56066377, -18.93426482, -19.3, -19.65819371,
            -20.00915068, -20.35315755, -20.69048425, -21.02138529, -21.3461009, -21.66485809,
            -21.97787159, -22.28534473, -22.58747025, -22.88443097, -23.17640052, -23.46354391,
            -23.74601809, -24.02397248, -24.29754947, -24.56688477, -24.83210793, -25.09334261,
            -25.35070701, -25.60431414, -25.85427211, -26.10068445, -26.34365036, -26.58326491,
            -26.8196193, -27.05280104, -27.2828942, -27.50997951, -27.73413461, -27.95543416,
            -28.17394999, -28.38975126, -28.6029046, -28.8134742, -29.02152195, -29.22710753,
            -29.43028856, -29.63112061, -29.82965739, -30.02595077, -30.22005088, -30.4120062,
            -30.6018636, -30.78966845, -30.97546465, -31.15929472, -31.34119983, -31.52121988,
            -31.69939353, -31.87575829, -32.05035051, -32.22320546, -32.39435738, -32.56383949,
            -32.73168408, -32.89792247, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33,
            -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33,
            -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33,
            -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33,
            -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33,
            -33, -33, -33, -33, -33, -33, -33
        ]

        spec_line_10db = [
            -7, -7, -7, -7, -7, -7, -7, -7, -7, -7, -7, -7, -7, -7, -7, -7, -7, -7, -7, -7, -7,
            -7, -7, -7, -7.967950853, -8.649284425, -9.304901072, -9.93667176, -10.54627042,
            -11.13520069, -11.70481826, -12.25634964, -12.7909081, -13.30950719, -13.81307228,
            -14.30245054, -14.77841947, -15.24169437, -15.69293479, -16.13275016, -16.56170477,
            -16.98032212, -17.38908873, -17.78845757, -18.17885106, -18.56066377, -18.93426482,
            -19.3, -19.65819371, -20.00915068, -20.35315755, -20.69048425, -21.02138529,
            -21.3461009, -21.66485809, -21.97787159, -22.28534473, -22.58747025, -22.88443097,
            -23.17640052, -23.46354391, -23.74601809, -24.02397248, -24.29754947, -24.56688477,
            -24.83210793, -25.09334261, -25.35070701, -25.60431414, -25.85427211, -26.10068445,
            -26.34365036, -26.58326491, -26.8196193, -27.05280104, -27.2828942, -27.50997951,
            -27.73413461, -27.95543416, -28.17394999, -28.38975126, -28.6029046, -28.8134742,
            -29.02152195, -29.22710753, -29.43028856, -29.63112061, -29.82965739, -30.02595077,
            -30.22005088, -30.4120062, -30.6018636, -30.78966845, -30.97546465, -31.15929472,
            -31.34119983, -31.52121988, -31.69939353, -31.87575829, -32.05035051, -32.22320546,
            -32.39435738, -32.56383949, -32.73168408, -32.89792247, -33, -33, -33, -33, -33,
            -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33,
            -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33,
            -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33,
            -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33, -33,
            -33, -33, -33
        ]


        
        group_10db_and_10MHz = [trace for trace in traces if float(trace["bandwidth"]) == 10000000.0 and int(trace["attenuation_setting"]) == 10]
        self.create_10mghz_plot(group_10db_and_10MHz, self.base_tds_filepath, gain_step=10, spec_line=spec_line_10db)

        group_max_gain_and_10MHz = [trace for trace in traces if float(trace["bandwidth"]) == 10000000.0 and int(trace["attenuation_setting"]) != 10]
        self.create_10mghz_plot(group_max_gain_and_10MHz, self.base_tds_filepath, gain_step=group_max_gain_and_10MHz[0]["attenuation_setting"], spec_line=max_gain_spec_line)

        group_10db_200MHz = [trace for trace in traces if float(trace["bandwidth"]) == 200000000.0 and int(trace["attenuation_setting"]) == 10]
        self.create_200mghz_plot(group_10db_200MHz, self.base_tds_filepath, gain_step=group_10db_200MHz[0]["attenuation_setting"])

        group_max_gain_200MHz = [trace for trace in traces if float(trace["bandwidth"]) == 200000000.0 and int(trace["attenuation_setting"]) != 10]
        self.create_200mghz_plot(group_max_gain_200MHz, self.base_tds_filepath, gain_step=group_max_gain_200MHz[0]["attenuation_setting"])

        group_200MHz = [trace for trace in traces if float(trace["bandwidth"]) == 200000000.0]

        group_wideband = [trace for trace in traces if float(trace["bandwidth"]) > 200000000.0]
        self.create_wideband_plot(group_wideband, self.base_tds_filepath, gain_step=group_wideband[0]["attenuation_setting"])

        def write_trace_data_to_xlsx(group, xlsx_filename):
            xlsx_filepath = os.path.join(self.base_tds_filepath, xlsx_filename)
            template = os.path.join(self.base_xlsx_filepath, xlsx_filename)

            if not os.path.exists(xlsx_filepath):
                # Use shutil to copy the file
                shutil.copyfile(template, xlsx_filepath)
            
            workbook = openpyxl.load_workbook(xlsx_filepath)
            sheet = workbook["raw data"]
            traces = []
            for i, trace in enumerate(group):
                # Transpose the trace data and take the values
                traces.append(trace['data'])

            trace_data = np.array(traces)
            trace_data = trace_data.T.tolist()
            for i, row in enumerate(trace_data):
                for j, value in enumerate(row):
                    cell = sheet.cell(row=i + 2, column=j + 2 + temp_coords_offset)
                    cell.value = value

            workbook.save(xlsx_filepath)
            workbook.close()

        write_trace_data_to_xlsx(group_10db_and_10MHz, xlsx_filepaths[0])
        write_trace_data_to_xlsx(group_max_gain_and_10MHz, xlsx_filepaths[1])
        write_trace_data_to_xlsx(group_200MHz, xlsx_filepaths[2])
        write_trace_data_to_xlsx(group_wideband, xlsx_filepaths[3])

    def write_gain_phase_tds(self, csv_filepath: str, temp: str, gain_phase: str):
        """
        Write gain phase data to an xlsx file.
        """

        # Get the xlsx filepath 
        base_filename = os.path.basename(csv_filepath)
        sn = base_filename.split("PATH")[1]
        if sn.startswith("1"):
            sn = "sn1"
        elif sn.startswith("2"):
            sn = "sn2"

        base_filename = base_filename.replace(".csv", "")
        xlsx_filename = self.tds_lookup.get(base_filename, {})
        xlsx_filename = xlsx_filename.get("filepaths", "")
        xlsx_filename = xlsx_filename + f" {temp.lower()}.xlsx"
        template_filepath = os.path.join(self.base_xlsx_filepath, xlsx_filename)
        xlsx_filepath = os.path.join(self.base_tds_filepath, xlsx_filename)

        if not os.path.exists(xlsx_filepath):
            # Use shutil to copy the file
            shutil.copyfile(template_filepath, xlsx_filepath)
        # Open the copied xlsx file
        workbook = openpyxl.load_workbook(xlsx_filepath)

        sheet = workbook[gain_phase + " " + sn]

        pd_data = pd.read_csv(csv_filepath, header=None)
        np_data = np.array(pd_data)

        # Create a list of dictionaries for metadata and raw data
        data = []
        frequencies = np_data[0][7:]  # First row is frequencie
        for i, row in enumerate(np_data[1:]):
            data.append({
                "attenuation_setting": row[0],
                "data": row[7:]
            })

        self.create_gain_phase_plot(
            data=data,
            frequencies=frequencies,
            destination_dir=self.base_tds_filepath,
            gain_phase=gain_phase
        )

        np_data = np_data.T.tolist()
        np_trace_data = np_data[7:]

        # Write the
        for i, row in enumerate(np_trace_data):
            for j, value in enumerate(row):
                if j == 0:
                    continue
                cell = sheet.cell(row=i + 2, column=j + 1)
                cell.value = float(value)

        # Save the workbook
        workbook.save(xlsx_filepath)
        workbook.close()


    def create_gain_phase_plot(self, data: List, frequencies: List, destination_dir: str, gain_phase: int):
        """
        Create and write gain or phase measurement plots
        """
        plt.figure(figsize=(10, 6))
        plt.title(gain_phase)
        plt.xlabel("Frequency (Hz)")
        plt.ylabel(gain_phase)
        plt.grid(True)

        frequencies = [float(freq) for freq in frequencies]
        for i, trace in enumerate(data):
            plt.plot(frequencies, [float(data) for data in trace["data"]], label=f"{trace['attenuation_setting']} dB")

        plt.legend()
        plt.savefig(os.path.join(destination_dir, f"{gain_phase}_plot.png"))
        plt.close()

    def write_s11_s22_tds(self, data: List, csv_filepath: str, temp: str, s11_s22: str):
        """
        Write S11 or S22 data to an xlsx file.
        """

                # Get the xlsx filepath 
        base_filename = os.path.basename(csv_filepath)
        sn = base_filename.split("PATH")[1]
        if sn.startswith("1"):
            sn = "sn1"
        elif sn.startswith("2"):
            sn = "sn2"

        base_filename = base_filename.replace(".csv", "")
        xlsx_filename = self.tds_lookup.get(base_filename, {})
        xlsx_filename = xlsx_filename.get("filepaths", "")
        xlsx_filename = xlsx_filename + f" {temp.lower()}.xlsx"
        template_filepath = os.path.join(self.base_xlsx_filepath, xlsx_filename)
        xlsx_filepath = os.path.join(self.base_tds_filepath, xlsx_filename)

        if not os.path.exists(xlsx_filepath):
            # Use shutil to copy the file
            shutil.copyfile(template_filepath, xlsx_filepath)

        # Open the copied xlsx file
        workbook = openpyxl.load_workbook(xlsx_filepath)

        sheet = workbook["vswr"]

        if s11_s22 == "S11":
            if sn == "sn1":
                col_coords = 2
            elif sn == "sn2":
                col_coords = 4
        elif s11_s22 == "S22":
            if sn == "sn1":
                col_coords = 3
            elif sn == "sn2":
                col_coords = 5

        np_data = np.array(data)

        # Create a list of dictionaries for metadata and raw data
        temp_data = []
        frequencies = np_data[0][7:]  # First row is frequencie
        for i, row in enumerate(np_data[1:]):
            temp_data.append({
                "attenuation_setting": row[0],
                "data": row[7:]
            })
        self.create_gain_phase_plot(
            data=temp_data,
            frequencies=frequencies,
            destination_dir=self.base_tds_filepath,
            gain_phase=s11_s22
        )

        np_data = np_data.T.tolist()
        np_trace_data = np_data[7:]

        # Write the data
        for i, row in enumerate(np_trace_data):
            for j, value in enumerate(row):
                if j == 0:
                    continue
                cell = sheet.cell(row=i + 2, column=col_coords)
                cell.value = float(value)

        # Save the workbook
        workbook.save(xlsx_filepath)
        workbook.close()

    def write_power_meter_tds(self, data: List, xlsx_template_filepath: str, new_xlsx_filepath: str, temp: str, sn: str):
        """
        Write power meter data to an xlsx file.
        """
        if not os.path.exists(new_xlsx_filepath):
            # Use shutil to copy the file
            shutil.copyfile(xlsx_template_filepath, new_xlsx_filepath)

        # Open the copied xlsx file
        workbook = openpyxl.load_workbook(new_xlsx_filepath)

        sheet = workbook[sn.upper() + " " + temp]

        # Write the data
        for i, row in enumerate(data):
            for j, value in enumerate(row):
                cell = sheet.cell(row=i + 1, column=j + 1)
                try:
                    cell.value = float(value)
                except ValueError:
                    cell.value = value

        # Save the workbook
        workbook.save(new_xlsx_filepath)
        workbook.close()

    def ask_temperature(self):
        root = tk.Tk()
        root.withdraw()  # Hide the main tkinter window

        temp_options = ["25c", "-13c", "55c"]
        temp_var = tk.StringVar(root)
        temp_var.set(temp_options[0])  # Set default value

        def on_select():
            root.quit()

        popup = tk.Toplevel(root)
        popup.title("Select Temperature")
        tk.Label(popup, text="Select the temperature:").pack(pady=10)
        dropdown = tk.OptionMenu(popup, temp_var, *temp_options)
        dropdown.pack(pady=10)

        add_to_existing_var = tk.BooleanVar()
        add_to_existing_check = tk.Checkbutton(popup, text="Add to existing sheet", variable=add_to_existing_var)
        add_to_existing_check.pack(pady=5)

        tk.Button(popup, text="OK", command=on_select).pack(pady=10)

        root.mainloop()
        selected_temp = temp_var.get()
        add_to_existing = add_to_existing_var.get()
        popup.destroy()
        root.destroy()

        return selected_temp, add_to_existing

    def run(self, data_dir: str, selected_temp: str):
        """
        Run the TestDataSheetWriter.
        """
        csv_filepaths = self.get_csv_filepaths(data_dir)
        # Iterate over the CSV filepaths and process each file
        for csv_filepath in csv_filepaths:
            # Get the corresponding xlsx filepath from the lookup table
            basename = os.path.basename(csv_filepath)

            if "BANDWIDTH" in basename:
                data = self.read_csv_file(csv_filepath)
                self.write_bandwidth_tds(data, csv_filepath, selected_temp)

            if "S21" in basename and "MLOG" in basename:
                data = self.read_csv_file(csv_filepath)
                self.write_gain_phase_tds(csv_filepath, selected_temp, "gain")

            if "S21" in basename and "PHASE" in basename:
                data = self.read_csv_file(csv_filepath)
                self.write_gain_phase_tds(csv_filepath, selected_temp, "phase")

            if "S11" in basename and "MLOG" in basename:
                data = self.read_csv_file(csv_filepath)
                self.write_s11_s22_tds(data, csv_filepath, selected_temp, "S11")

            if "S22" in basename and "MLOG" in basename:
                data = self.read_csv_file(csv_filepath)
                self.write_s11_s22_tds(data, csv_filepath, selected_temp, "S22")

if __name__ == "__main__":
    # Create an instance of the TestDataSheetWriter class
    writer = TestDataSheetWriter()
    temp, add_to_exsiting = writer.ask_temperature()
    sn1_data_dir = writer.get_data_dir()
    if add_to_exsiting:
        writer.ask_user_for_write_dir()
    # Run the writer for both directories
    writer.run(sn1_data_dir, temp)
    print("Test data sheets have been written successfully.")